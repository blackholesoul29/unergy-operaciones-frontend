"""
scripts/migrate_fallas.py
Migración de 1.081 fallas desde el Excel histórico a PostgreSQL.

Uso:
  pip install openpyxl
  EXCEL_PATH=/path/to/Identificacion_de_falla.xlsx python scripts/migrate_fallas.py

Variables de entorno:
  EXCEL_PATH   — ruta al archivo Excel (requerido)
  DATABASE_URL — si quieres sobreescribir la URL de la DB
  DRY_RUN      — si "1", solo muestra el resumen sin insertar

Lógica de matching de proyectos:
  1. Busca por nombre_clientes exacto
  2. Busca por nombre_display exacto
  3. Busca por nombre_bitacora exacto
  4. Si no encuentra, guarda el nombre en proyecto_nombre_raw (campo texto)
     para que el equipo lo mapee manualmente después.
"""
import sys
import os
import warnings
warnings.filterwarnings("ignore")

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

from app.core.database import SessionLocal
from app.models.fallas import (
    Falla, FallaSeguimiento, FallaFoto,
    FallaCatTipo, FallaCatEstado, FallaCatPrioridad, FallaCatResolucion,
)

EXCEL_PATH = os.getenv("EXCEL_PATH", "Identificacion_de_falla.xlsx")
DRY_RUN    = os.getenv("DRY_RUN", "0") == "1"


# ─── helpers ────────────────────────────────────────────────

def _norm(s) -> str:
    return str(s).strip().lower() if s else ""


def _safe_str(v) -> str | None:
    if v is None: return None
    s = str(v).strip()
    return s if s else None


def _safe_dt(v) -> datetime | None:
    if v is None: return None
    if isinstance(v, datetime): return v
    return None


def _next_codigo(db, counter: list) -> str:
    counter[0] += 1
    return f"F-{counter[0]:05d}"


def _build_project_map(db):
    """Construye un mapa nombre→proyecto_id buscando en todos los campos de nombre"""
    from sqlalchemy import text
    rows = db.execute(text(
        "SELECT id, sub_project, nombre_display, nombre_bitacora, nombre_clientes "
        "FROM proyectos WHERE id IS NOT NULL"
    )).fetchall()
    m = {}
    for r in rows:
        pid = str(r.id)
        for alias in [r.sub_project, r.nombre_display, r.nombre_bitacora, r.nombre_clientes]:
            if alias:
                m[_norm(alias)] = pid
    return m


def _build_cat_maps(db):
    """Construye mapas para tipos, estados, prioridades y resoluciones"""
    tipos = {_norm(t.codigo): str(t.id) for t in db.query(FallaCatTipo).all()}
    estados_map = {}
    for e in db.query(FallaCatEstado).all():
        estados_map[_norm(e.codigo)]   = str(e.id)
        estados_map[_norm(e.etiqueta)] = str(e.id)
    prios = {}
    for p in db.query(FallaCatPrioridad).all():
        prios[_norm(p.codigo)]   = str(p.id)
        prios[_norm(p.etiqueta)] = str(p.id)
    resols = {}
    for r in db.query(FallaCatResolucion).all():
        resols[_norm(r.etiqueta)] = str(r.id)
        resols[_norm(r.codigo)]   = str(r.id)
    return tipos, estados_map, prios, resols


def _estado_code(raw: str) -> str:
    mapping = {
        "activa":      "activa",
        "en revisión": "en_revision",
        "en revision": "en_revision",
        "programada":  "programada",
        "terminada":   "terminada",
    }
    return mapping.get(_norm(raw), _norm(raw))


# ─── main ───────────────────────────────────────────────────

def migrate():
    if not Path(EXCEL_PATH).exists():
        print(f"❌ No se encontró el archivo: {EXCEL_PATH}")
        sys.exit(1)

    print(f"📖 Leyendo: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb["Fallas"]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]: rows.append(row)
    print(f"   {len(rows)} fallas encontradas en el Excel")

    if DRY_RUN:
        print("🔍 DRY RUN — no se insertará nada en la DB")

    db = SessionLocal()
    try:
        proj_map  = _build_project_map(db)
        tipo_map, estado_map, prio_map, resol_map = _build_cat_maps(db)

        # Obtener el máximo código actual para continuar la secuencia
        from sqlalchemy import func as sqlfunc
        last_code = db.query(sqlfunc.max(Falla.codigo_interno)).scalar()
        if last_code:
            try:    start = int(last_code.split("-")[1])
            except: start = 0
        else:
            start = 0
        counter = [start]

        stats = {
            "insertadas": 0, "duplicadas": 0, "sin_proyecto": 0,
            "seguimientos": 0, "fotos": 0, "errores": 0
        }
        no_match_projs = set()

        # Índice de códigos existentes para evitar duplicados
        existing = {c for (c,) in db.query(Falla.codigo_interno).all()}

        BATCH = 100
        batch_fallas = []
        batch_segs   = []
        batch_fotos  = []

        for i, row in enumerate(rows, 1):
            try:
                codigo_orig = _safe_str(row[0])   # F-00001
                proj_name   = _safe_str(row[1])   # "La Paz Verso"
                falla_code  = _safe_str(row[2])   # "2.8"
                desc_falla  = _safe_str(row[3])   # "Falla de inversor"
                # row[4] = categoría (se deriva del tipo)
                estado_raw  = _safe_str(row[5])
                fecha_ident = row[6]
                hora_ident  = row[7]
                fecha_ocurr = _safe_dt(row[8])
                resol_raw   = _safe_str(row[9])
                prio_raw    = _safe_str(row[10])
                descripcion = _safe_str(row[11])
                seguimiento = _safe_str(row[12])
                foto_link   = _safe_str(row[13])
                fecha_term  = _safe_dt(row[14])
                centinela   = _safe_str(row[15])
                fecha_reg   = _safe_dt(row[16])
                ultima_act  = _safe_dt(row[17])

                # Evitar duplicados si se vuelve a correr la migración
                if codigo_orig and codigo_orig in existing:
                    stats["duplicadas"] += 1
                    continue

                # Match de proyecto
                proj_id    = proj_map.get(_norm(proj_name)) if proj_name else None
                proj_raw   = proj_name if not proj_id else None
                if not proj_id and proj_name:
                    no_match_projs.add(proj_name)
                    stats["sin_proyecto"] += 1

                # IDs de catálogos
                tipo_id   = tipo_map.get(_norm(str(falla_code))) if falla_code else None
                estado_cd = _estado_code(estado_raw or "")
                estado_id = estado_map.get(estado_cd)
                prio_id   = prio_map.get(_norm(prio_raw)) if prio_raw else None
                resol_id  = resol_map.get(_norm(resol_raw)) if resol_raw else None

                # Fecha registro
                if fecha_reg and isinstance(fecha_reg, datetime):
                    f_registro = fecha_reg
                elif fecha_ident and isinstance(fecha_ident, datetime):
                    f_registro = fecha_ident
                else:
                    f_registro = datetime.utcnow()

                # Días abierta y SLA
                ref_fecha = fecha_term or ultima_act or datetime.utcnow()
                dias_ab   = (ref_fecha - f_registro).days if isinstance(ref_fecha, datetime) else None

                # Usar código original del Excel, si ya hay conflicto generar nuevo
                codigo_final = codigo_orig or _next_codigo(db, counter)

                falla = Falla(
                    codigo_interno       = codigo_final,
                    proyecto_id          = proj_id,
                    proyecto_nombre_raw  = proj_raw,
                    tipo_falla_id        = tipo_id,
                    estado_id            = estado_id,
                    prioridad_id         = prio_id,
                    tipo_resolucion_id   = resol_id,
                    descripcion          = descripcion or desc_falla,
                    fecha_identificacion = fecha_ident.date() if isinstance(fecha_ident, datetime) else (
                        fecha_ident if hasattr(fecha_ident, "year") else None),
                    hora_identificacion  = hora_ident if hasattr(hora_ident, "hour") else None,
                    fecha_ocurrencia     = fecha_ocurr,
                    fecha_resolucion     = fecha_term,
                    centinela            = centinela,
                    fecha_registro       = f_registro,
                    ultima_actualizacion = ultima_act,
                    dias_abierta         = dias_ab,
                    sla_cumplido         = None,  # se recalcula vía endpoint
                )
                batch_fallas.append(falla)

                # Seguimiento
                if seguimiento:
                    batch_segs.append((codigo_final, seguimiento, f_registro))
                    stats["seguimientos"] += 1

                # Foto Drive
                if foto_link:
                    batch_fotos.append((codigo_final, foto_link))
                    stats["fotos"] += 1

                stats["insertadas"] += 1

                # Insertar en lotes de BATCH
                if len(batch_fallas) >= BATCH and not DRY_RUN:
                    db.bulk_save_objects(batch_fallas)
                    db.flush()
                    _insert_seguimientos(db, batch_fallas, batch_segs)
                    _insert_fotos(db, batch_fallas, batch_fotos)
                    db.commit()
                    batch_fallas.clear()
                    batch_segs.clear()
                    batch_fotos.clear()
                    print(f"   → {stats['insertadas']} fallas insertadas…", end="\r")

            except Exception as e:
                stats["errores"] += 1
                print(f"\n⚠️  Error en fila {i}: {e}")

        # Último lote
        if batch_fallas and not DRY_RUN:
            db.bulk_save_objects(batch_fallas)
            db.flush()
            _insert_seguimientos(db, batch_fallas, batch_segs)
            _insert_fotos(db, batch_fallas, batch_fotos)
            db.commit()

        # Reporte
        print(f"\n{'─'*50}")
        print(f"✅ Fallas insertadas:     {stats['insertadas']}")
        print(f"⏭️  Duplicadas (omitidas): {stats['duplicadas']}")
        print(f"📝 Sin match de proyecto: {stats['sin_proyecto']}")
        print(f"💬 Seguimientos creados:  {stats['seguimientos']}")
        print(f"🖼️  Links de fotos Drive:  {stats['fotos']}")
        print(f"❌ Errores:               {stats['errores']}")
        if no_match_projs:
            print(f"\n⚠️  Proyectos sin match ({len(no_match_projs)}) — agregar a tabla 'proyectos':")
            for p in sorted(no_match_projs):
                print(f"   • {p}")
        print(f"{'─'*50}")
        if DRY_RUN:
            print("🔍 DRY RUN — nada fue insertado.")

    except Exception as e:
        db.rollback()
        print(f"\n❌ Error crítico: {e}")
        raise
    finally:
        db.close()


def _insert_seguimientos(db, fallas_batch: list, segs: list):
    """Inserta seguimientos usando el código_interno como clave"""
    code_to_id = {f.codigo_interno: f.id for f in fallas_batch}
    for (codigo, nota, fecha) in segs:
        fid = code_to_id.get(codigo)
        # Si la falla ya estaba en DB antes de este batch, buscar por código
        if not fid:
            existing = db.query(Falla.id).filter(
                Falla.codigo_interno == codigo).scalar()
            fid = existing
        if fid:
            db.add(FallaSeguimiento(falla_id=fid, nota=nota, created_at=fecha))


def _insert_fotos(db, fallas_batch: list, fotos: list):
    """Inserta links de Drive como FallaFoto"""
    code_to_id = {f.codigo_interno: f.id for f in fallas_batch}
    for (codigo, url) in fotos:
        fid = code_to_id.get(codigo)
        if not fid:
            existing = db.query(Falla.id).filter(
                Falla.codigo_interno == codigo).scalar()
            fid = existing
        if fid:
            db.add(FallaFoto(falla_id=fid, url=url, etapa="historico", orden=0))


if __name__ == "__main__":
    migrate()
